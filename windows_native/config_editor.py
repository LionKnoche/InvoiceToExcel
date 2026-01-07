"""
Invoice Config Editor - GUI zur Konfiguration von Kategorien, Spalten und Vendors.

Kategorien definieren Sheet-Schemata (Spaltenlisten).
Vendors werden Kategorien zugewiesen und können pro Spalte einen spezifischen Prompt haben.

Pipeline-Tab ermöglicht Human-in-the-Loop Workflow:
1. Dateien hinzufügen
2. OCR + Regex-Klassifikation
3. User überprüft/korrigiert Vendor-Zuordnung
4. LLM-Extraktion starten
"""
import os
import re
import threading
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import messagebox, simpledialog, ttk, filedialog


# Import config_store/pipeline (funktioniert sowohl als Package als auch standalone)
try:
    from . import config_store
    from . import pipeline
except ImportError:
    _module_dir = Path(__file__).parent
    if str(_module_dir) not in sys.path:
        sys.path.insert(0, str(_module_dir))
    import config_store
    import pipeline

CONFIG_PATH = config_store.CONFIG_PATH

# Vordefinierte Spalten mit Default-Prompts für das LLM
PREDEFINED_COLUMNS: Dict[str, str] = {
    "InvoiceId": "Automatisch generierte ID (nicht aus Rechnung extrahieren).",
    "InvoiceNumber": "Extrahiere die Rechnungsnummer (z.B. 'Rechnungs-Nr.', 'Invoice No.').",
    "VendorName": "Extrahiere den Namen des Lieferanten/Rechnungsstellers.",
    "InvoiceDate": "Extrahiere das Rechnungsdatum im Format YYYY-MM-DD.",
    "DueDate": "Extrahiere das Fälligkeitsdatum/Zahlungsziel im Format YYYY-MM-DD.",
    "GrossAmount": "Extrahiere den Bruttobetrag (Gesamtsumme inkl. MwSt.).",
    "NetAmount": "Extrahiere den Nettobetrag (ohne MwSt.).",
    "TaxAmount": "Extrahiere den Steuerbetrag (MwSt.-Betrag).",
    "TaxRate": "Extrahiere den Steuersatz als Dezimalzahl (z.B. 0.19 für 19%).",
    "Currency": "Extrahiere die Währung (z.B. EUR, USD).",
    "Description": "Extrahiere eine kurze Beschreibung der Leistung/des Produkts.",
    "IBAN": "Extrahiere die IBAN-Bankverbindung.",
    "BIC": "Extrahiere den BIC/SWIFT-Code.",
    "PaymentTerms": "Extrahiere die Zahlungsbedingungen.",
    "PurchaseOrderNumber": "Extrahiere die Bestellnummer (falls vorhanden).",
    "CustomerNumber": "Extrahiere die Kundennummer.",
    "FilePath": "Pfad zur Rechnungsdatei (automatisch gesetzt).",
}

DEFAULT_LLM_CONFIG: Dict[str, str] = config_store.DEFAULT_CONFIG["llm"].copy()


def load_config() -> Dict[str, Any]:
    def on_error(message: str) -> None:
        messagebox.showwarning("Warnung", message)

    return config_store.load_config(on_error=on_error)


def save_config(cfg: Dict[str, Any]) -> None:
    try:
        config_store.save_config(cfg)
    except Exception as e:
        messagebox.showerror("Fehler", f"Konfiguration konnte nicht gespeichert werden: {e}")


def get_default_column_prompt(column_name: str) -> str:
    """Gibt den Default-Prompt für eine Spalte zurück."""
    return PREDEFINED_COLUMNS.get(column_name, f"Extrahiere den Wert für '{column_name}'.")


class ConfigEditor(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Invoice-Konfiguration")
        self.geometry("1100x750")

        self.cfg: Dict[str, Any] = load_config()
        self.llm_provider_var = tk.StringVar()
        self.llm_base_url_var = tk.StringVar()
        self.llm_model_var = tk.StringVar()
        self.llm_api_key_var = tk.StringVar()

        notebook = ttk.Notebook(self)
        notebook.pack(fill=tk.BOTH, expand=True)

        pipeline_frame = tk.Frame(notebook)
        vendor_frame = tk.Frame(notebook)
        category_frame = tk.Frame(notebook)
        global_frame = tk.Frame(notebook)

        notebook.add(pipeline_frame, text="Pipeline")
        notebook.add(vendor_frame, text="Vendors")
        notebook.add(category_frame, text="Kategorien/Sheets")
        notebook.add(global_frame, text="Globaler Prompt")

        self._build_pipeline_tab(pipeline_frame)
        self._build_vendor_tab(vendor_frame)
        self._build_category_tab(category_frame)
        self._build_global_tab(global_frame)

        # Pipeline-Daten: Liste von {"file_path": Path, "ocr_text": str, "detected_vendor": str, "selected_vendor": str, "status": str}
        self.pipeline_items: List[Dict[str, Any]] = []

        # Intern gespeicherter Index für aktuell ausgewählten Vendor/Kategorie
        self._current_vendor_idx: int = -1
        self._current_category_idx: int = -1

        self.refresh_category_list()
        self.refresh_vendor_list()
        self.load_default_prompt()
        self.load_llm_settings()
        self.refresh_vendor_category_choices()
        self.pipeline_refresh_vendor_choices()

    # ========== Pipeline-Tab ==========
    def _build_pipeline_tab(self, parent: tk.Frame) -> None:
        """Pipeline-Tab: Vereinfachter Workflow für Rechnungsverarbeitung."""
        # Oberer Bereich: Buttons
        top_frame = tk.Frame(parent)
        top_frame.pack(fill=tk.X, padx=5, pady=5)

        tk.Button(top_frame, text="Dateien hinzufügen...", command=self.pipeline_add_files).pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="Ausgewählte entfernen", command=self.pipeline_remove_selected).pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="Alle entfernen", command=self.pipeline_clear_all).pack(side=tk.LEFT, padx=5)

        tk.Label(top_frame, text="   ").pack(side=tk.LEFT)  # Spacer
        tk.Button(top_frame, text="Extraktion starten", command=self.pipeline_run_extraction, bg="#2196F3", fg="white").pack(side=tk.LEFT, padx=5)

        # Hinweis
        hint_frame = tk.Frame(parent)
        hint_frame.pack(fill=tk.X, padx=5)
        tk.Label(
            hint_frame,
            text="Workflow: 1) Dateien hinzufügen → 2) Vendor auswählen → 3) Extraktion starten",
            fg="gray",
        ).pack(anchor="w")

        # Hauptbereich: Treeview mit Dateien
        tree_frame = tk.Frame(parent)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        columns = ("file", "status", "selected_vendor", "category", "sheet_name")
        self.pipeline_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="extended")
        self.pipeline_tree.heading("file", text="Datei")
        self.pipeline_tree.heading("status", text="Status")
        self.pipeline_tree.heading("selected_vendor", text="Vendor")
        self.pipeline_tree.heading("category", text="Kategorie")
        self.pipeline_tree.heading("sheet_name", text="Excel-Sheet")

        self.pipeline_tree.column("file", width=250)
        self.pipeline_tree.column("status", width=120)
        self.pipeline_tree.column("selected_vendor", width=120)
        self.pipeline_tree.column("category", width=120)
        self.pipeline_tree.column("sheet_name", width=120)

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.pipeline_tree.yview)
        self.pipeline_tree.configure(yscrollcommand=scrollbar.set)

        self.pipeline_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Doppelklick zum Ändern des Vendors
        self.pipeline_tree.bind("<Double-1>", self.pipeline_on_double_click)

        # Unterer Bereich: Vendor-Auswahl für markierte Einträge
        bottom_frame = tk.Frame(parent)
        bottom_frame.pack(fill=tk.X, padx=5, pady=5)

        tk.Label(bottom_frame, text="Vendor für markierte Einträge setzen:").pack(side=tk.LEFT)
        self.pipeline_vendor_var = tk.StringVar()
        self.pipeline_vendor_combo = ttk.Combobox(bottom_frame, textvariable=self.pipeline_vendor_var, state="readonly", width=25)
        self.pipeline_vendor_combo.pack(side=tk.LEFT, padx=5)
        tk.Button(bottom_frame, text="Anwenden", command=self.pipeline_apply_vendor).pack(side=tk.LEFT, padx=5)

        # Status-Label mit Log-Button
        status_frame = tk.Frame(parent)
        status_frame.pack(fill=tk.X, padx=5, pady=2)
        
        self.pipeline_status_label = tk.Label(status_frame, text="Bereit.", anchor="w")
        self.pipeline_status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Log-Datei öffnen Button
        log_path = config_store.get_log_path()
        tk.Button(
            status_frame, 
            text="📋 Log öffnen", 
            command=lambda: self._open_log_file(log_path),
            font=("TkDefaultFont", 8)
        ).pack(side=tk.RIGHT, padx=(5, 0))

    def _open_log_file(self, log_path: Path) -> None:
        """Öffnet die Log-Datei im Standard-Texteditor."""
        import subprocess
        import platform
        
        if not log_path.exists():
            messagebox.showinfo("Info", "Log-Datei existiert noch nicht.\nFühre zuerst eine Extraktion aus.")
            return
        
        try:
            if platform.system() == "Windows":
                os.startfile(str(log_path))
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", str(log_path)])
            else:  # Linux
                subprocess.run(["xdg-open", str(log_path)])
        except Exception as e:
            messagebox.showerror("Fehler", f"Konnte Log-Datei nicht öffnen:\n{e}")

    def pipeline_refresh_vendor_choices(self) -> None:
        """Aktualisiert die Vendor-Dropdown-Liste in der Pipeline."""
        names = [v.get("name", "") for v in self.cfg.get("vendors", [])]
        self.pipeline_vendor_combo["values"] = ["(Unbekannt)"] + names

    def pipeline_add_files(self) -> None:
        """Öffnet Dateidialog zum Hinzufügen von Rechnungsdateien."""
        filetypes = [
            ("PDF und Bilder", "*.pdf *.PDF *.png *.PNG *.jpg *.JPG *.jpeg *.JPEG"),
            ("PDF-Dateien", "*.pdf *.PDF"),
            ("Bilddateien", "*.png *.PNG *.jpg *.JPG *.jpeg *.JPEG"),
            ("Alle Dateien", "*.*"),
        ]
        files = filedialog.askopenfilenames(title="Rechnungsdateien auswählen", filetypes=filetypes)
        if not files:
            return

        for file_path in files:
            path = Path(file_path)
            # Prüfen, ob schon vorhanden
            if any(item["file_path"] == path for item in self.pipeline_items):
                continue
            item = {
                "file_path": path,
                "ocr_text": "",
                "selected_vendor": "(Bitte auswählen)",
                "status": "Bereit",
            }
            self.pipeline_items.append(item)

        self._pipeline_refresh_tree()
        self.pipeline_refresh_vendor_choices()

    def pipeline_remove_selected(self) -> None:
        """Entfernt die ausgewählten Einträge aus der Pipeline."""
        selected = self.pipeline_tree.selection()
        if not selected:
            return
        indices_to_remove = [int(iid) for iid in selected]
        self.pipeline_items = [item for i, item in enumerate(self.pipeline_items) if i not in indices_to_remove]
        self._pipeline_refresh_tree()

    def pipeline_clear_all(self) -> None:
        """Entfernt alle Einträge aus der Pipeline."""
        if self.pipeline_items and not messagebox.askyesno("Bestätigen", "Alle Einträge entfernen?"):
            return
        self.pipeline_items.clear()
        self._pipeline_refresh_tree()

    def _pipeline_refresh_tree(self) -> None:
        """Aktualisiert die Treeview-Anzeige mit Kategorie und Sheet-Info."""
        self.pipeline_tree.delete(*self.pipeline_tree.get_children())
        for i, item in enumerate(self.pipeline_items):
            vendor_name = item["selected_vendor"]
            category_name = "-"
            sheet_name = "-"
            
            # Kategorie und Sheet-Name aus Vendor-Config ermitteln
            if vendor_name and vendor_name not in ("(Bitte auswählen)", "(Unbekannt)", ""):
                for vendor in self.cfg.get("vendors", []):
                    if vendor.get("name") == vendor_name:
                        category_name = vendor.get("category", "-") or "-"
                        # Sheet-Name aus Kategorie holen
                        for cat in self.cfg.get("categories", []):
                            if cat.get("name") == category_name:
                                sheet_name = cat.get("sheet_name", "-") or "-"
                                break
                        break
            
            self.pipeline_tree.insert(
                "",
                "end",
                iid=str(i),
                values=(
                    item["file_path"].name,
                    item["status"],
                    vendor_name,
                    category_name,
                    sheet_name,
                ),
            )

    def pipeline_on_double_click(self, event: Any) -> None:
        """Doppelklick auf einen Eintrag: Vendor ändern."""
        item_id = self.pipeline_tree.identify_row(event.y)
        if not item_id:
            return
        idx = int(item_id)
        if idx >= len(self.pipeline_items):
            return

        # Dialog zur Vendor-Auswahl
        vendors = ["(Unbekannt)"] + [v.get("name", "") for v in self.cfg.get("vendors", [])]
        current = self.pipeline_items[idx]["selected_vendor"]

        dialog = tk.Toplevel(self)
        dialog.title("Vendor auswählen")
        dialog.geometry("300x100")
        dialog.transient(self)
        dialog.grab_set()

        tk.Label(dialog, text=f"Vendor für: {self.pipeline_items[idx]['file_path'].name}").pack(pady=5)
        var = tk.StringVar(value=current)
        combo = ttk.Combobox(dialog, textvariable=var, values=vendors, state="readonly", width=30)
        combo.pack(pady=5)

        def on_ok():
            self.pipeline_items[idx]["selected_vendor"] = var.get()
            self._pipeline_refresh_tree()
            dialog.destroy()

        tk.Button(dialog, text="OK", command=on_ok).pack(pady=5)

    def pipeline_apply_vendor(self) -> None:
        """Wendet den gewählten Vendor auf alle markierten Einträge an."""
        selected = self.pipeline_tree.selection()
        vendor = self.pipeline_vendor_var.get()
        if not selected or not vendor:
            return
        for iid in selected:
            idx = int(iid)
            if idx < len(self.pipeline_items):
                self.pipeline_items[idx]["selected_vendor"] = vendor
        self._pipeline_refresh_tree()

    def pipeline_run_ocr(self) -> None:
        """Führt OCR und Regex-Klassifikation auf allen Dateien mit Status 'Hinzugefügt' aus."""
        items_to_process = [item for item in self.pipeline_items if item["status"] == "Hinzugefügt"]
        if not items_to_process:
            messagebox.showinfo("Info", "Keine neuen Dateien zum Verarbeiten.")
            return

        self.pipeline_status_label.config(text=f"OCR läuft für {len(items_to_process)} Datei(en)...")
        self.update()

        # In einem Thread ausführen, um UI nicht zu blockieren
        def run_ocr_thread():
            try:
                def update_status(value):
                    self.after(0, lambda v=value: self.pipeline_status_label.config(text=v))

                def refresh_tree():
                    self.after(0, self._pipeline_refresh_tree)

                pipeline.run_ocr(
                    items_to_process,
                    update_status=update_status,
                    refresh_tree=refresh_tree,
                )
            except Exception as e:
                err_msg = str(e)
                self.after(0, lambda msg=err_msg: self.pipeline_status_label.config(text=f"Fehler: {msg}"))

        thread = threading.Thread(target=run_ocr_thread, daemon=True)
        thread.start()

    def pipeline_run_extraction(self) -> None:
        """Führt OCR + LLM-Extraktion für bereite Einträge aus."""
        items_to_extract = [item for item in self.pipeline_items if item["status"] == "Bereit"]
        if not items_to_extract:
            messagebox.showinfo("Info", "Keine Dateien zum Extrahieren. Bitte zuerst Dateien hinzufügen.")
            return

        # Prüfen, ob alle einen Vendor haben
        unknown_items = [item for item in items_to_extract if item["selected_vendor"] in ("(Bitte auswählen)", "(Unbekannt)", "")]
        if unknown_items:
            messagebox.showwarning(
                "Vendor fehlt",
                f"{len(unknown_items)} Datei(en) haben keinen Vendor zugewiesen.\nBitte zuerst Vendor auswählen.",
            )
            return

        self.pipeline_status_label.config(text=f"LLM-Extraktion läuft für {len(items_to_extract)} Datei(en)...")
        self.update()

        def run_extraction_thread():
            try:
                def update_status(value):
                    self.after(0, lambda v=value: self.pipeline_status_label.config(text=v))

                def refresh_tree():
                    self.after(0, self._pipeline_refresh_tree)

                pipeline.run_extraction(
                    items_to_extract,
                    update_status=update_status,
                    refresh_tree=refresh_tree,
                )
            except Exception as e:
                err_msg = str(e)
                self.after(0, lambda msg=err_msg: self.pipeline_status_label.config(text=f"Fehler: {msg}"))

        thread = threading.Thread(target=run_extraction_thread, daemon=True)
        thread.start()

    # ========== Vendor-Tab ==========
    def _build_vendor_tab(self, parent: tk.Frame) -> None:
        # Linke Seite: Vendor-Liste
        left_frame = tk.Frame(parent)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)

        self.vendor_listbox = tk.Listbox(left_frame, width=20)
        self.vendor_listbox.pack(fill=tk.Y, expand=True)
        self.vendor_listbox.bind("<<ListboxSelect>>", self.on_vendor_select)

        btn_frame = tk.Frame(left_frame)
        btn_frame.pack(fill=tk.X, pady=5)
        tk.Button(btn_frame, text="Vendor hinzufügen", command=self.add_vendor).pack(fill=tk.X, pady=2)
        tk.Button(btn_frame, text="Vendor löschen", command=self.delete_vendor).pack(fill=tk.X, pady=2)

        # Rechte Seite: Details
        right_frame = tk.Frame(parent)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Oberer Bereich: Name, Kategorie, Patterns
        top_frame = tk.Frame(right_frame)
        top_frame.pack(fill=tk.X)

        tk.Label(top_frame, text="Vendor-Name:").grid(row=0, column=0, sticky="w")
        self.vendor_name_var = tk.StringVar()
        tk.Entry(top_frame, textvariable=self.vendor_name_var, width=30).grid(row=0, column=1, sticky="w", padx=5)

        tk.Label(top_frame, text="Kategorie (Sheet-Schema):").grid(row=1, column=0, sticky="w", pady=(5, 0))
        self.vendor_category_var = tk.StringVar()
        self.vendor_category_combo = ttk.Combobox(top_frame, textvariable=self.vendor_category_var, state="readonly", width=27)
        self.vendor_category_combo.grid(row=1, column=1, sticky="w", padx=5, pady=(5, 0))
        self.vendor_category_combo.bind("<<ComboboxSelected>>", self.on_vendor_category_change)

        tk.Label(top_frame, text="Allgemeiner Vendor-Prompt:").grid(row=2, column=0, sticky="nw", pady=(5, 0))
        self.vendor_prompt_text = tk.Text(top_frame, height=3, width=40)
        self.vendor_prompt_text.grid(row=2, column=1, sticky="w", padx=5, pady=(5, 0))

        # Unterer Bereich: Spalten-Prompts
        tk.Label(right_frame, text="Spalten-spezifische Prompts (leer = Default-Prompt):").pack(anchor="w", pady=(10, 0))

        columns_frame = tk.Frame(right_frame)
        columns_frame.pack(fill=tk.BOTH, expand=True)

        # Canvas mit Scrollbar für Spalten-Prompts
        canvas = tk.Canvas(columns_frame)
        scrollbar = ttk.Scrollbar(columns_frame, orient="vertical", command=canvas.yview)
        self.column_prompts_frame = tk.Frame(canvas)

        self.column_prompts_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.column_prompts_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Speichern: Dictionary von column_name -> Entry-Widget
        self.column_prompt_entries: Dict[str, tk.Text] = {}

        # Buttons
        btn_frame2 = tk.Frame(right_frame)
        btn_frame2.pack(fill=tk.X, pady=5)
        tk.Button(btn_frame2, text="Vendor-Eintrag speichern", command=self.save_current_vendor).pack(side=tk.LEFT, padx=5)

    def on_vendor_category_change(self, event: Any = None) -> None:
        """Wenn die Kategorie gewechselt wird, Spalten-Prompts neu aufbauen."""
        self._rebuild_column_prompts()

    def _rebuild_column_prompts(self, vendor_column_prompts: Optional[Dict[str, Any]] = None) -> None:
        """Baut die Spalten-Prompt-Eingabefelder basierend auf der gewählten Kategorie neu auf."""
        # Alte Widgets entfernen
        for widget in self.column_prompts_frame.winfo_children():
            widget.destroy()
        self.column_prompt_entries.clear()

        category_name = self.vendor_category_var.get()
        if not category_name:
            tk.Label(self.column_prompts_frame, text="(Bitte zuerst eine Kategorie wählen)").pack(anchor="w")
            return

        # Kategorie finden
        category = None
        for cat in self.cfg.get("categories", []):
            if cat.get("name") == category_name:
                category = cat
                break

        if not category:
            tk.Label(self.column_prompts_frame, text="(Kategorie nicht gefunden)").pack(anchor="w")
            return

        columns = category.get("columns", [])
        if not columns:
            tk.Label(self.column_prompts_frame, text="(Keine Spalten in dieser Kategorie)").pack(anchor="w")
            return

        vendor_column_prompts = vendor_column_prompts or {}
        # Parse vendor_column_prompts: kann Strings oder Objekte sein
        parsed_prompts: Dict[str, Any] = {}
        for col_name, col_config in vendor_column_prompts.items():
            if isinstance(col_config, str):
                parsed_prompts[col_name] = col_config
            elif isinstance(col_config, dict):
                # Nur den Prompt-Text verwenden, keine Lookup-Konfiguration
                parsed_prompts[col_name] = col_config.get("prompt", "")

        for col in columns:
            # Haupt-Frame für diese Spalte
            main_frame = tk.Frame(self.column_prompts_frame)
            main_frame.pack(fill=tk.X, pady=5)
            
            # Erste Zeile: Prompt-Eingabe
            prompt_frame = tk.Frame(main_frame)
            prompt_frame.pack(fill=tk.X, pady=2)

            default_prompt = get_default_column_prompt(col)
            tk.Label(prompt_frame, text=f"{col}:", width=18, anchor="w").pack(side=tk.LEFT)

            text_widget = tk.Text(prompt_frame, height=2, width=60)
            text_widget.pack(side=tk.LEFT, padx=5)

            # Wenn vendor-spezifischer Prompt existiert, diesen anzeigen
            col_config = parsed_prompts.get(col, {})
            vendor_prompt = col_config.get("prompt", "") if isinstance(col_config, dict) else (col_config if isinstance(col_config, str) else "")
            if vendor_prompt:
                text_widget.insert("1.0", vendor_prompt)

            # Label mit Default-Hint
            hint_label = tk.Label(prompt_frame, text=f"[Default: {default_prompt[:50]}...]", fg="gray", font=("TkDefaultFont", 8))
            hint_label.pack(side=tk.LEFT, padx=2)

            self.column_prompt_entries[col] = text_widget

    # ========== Kategorien-Tab ==========
    def _build_category_tab(self, parent: tk.Frame) -> None:
        # Linke Seite: Kategorie-Liste
        left_frame = tk.Frame(parent)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)

        self.category_listbox = tk.Listbox(left_frame, width=20)
        self.category_listbox.pack(fill=tk.Y, expand=True)
        self.category_listbox.bind("<<ListboxSelect>>", self.on_category_select)

        btn_frame = tk.Frame(left_frame)
        btn_frame.pack(fill=tk.X, pady=5)
        tk.Button(btn_frame, text="Kategorie hinzufügen", command=self.add_category).pack(fill=tk.X, pady=2)
        tk.Button(btn_frame, text="Kategorie löschen", command=self.delete_category).pack(fill=tk.X, pady=2)
        tk.Label(btn_frame, text="").pack(pady=2)  # Spacer
        tk.Button(btn_frame, text="📋 Excel importieren...", command=self.import_excel_categories, bg="#4CAF50", fg="white").pack(fill=tk.X, pady=2)

        # Rechte Seite: Details
        right_frame = tk.Frame(parent)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Name und Sheet-Name
        top_frame = tk.Frame(right_frame)
        top_frame.pack(fill=tk.X)

        tk.Label(top_frame, text="Kategorie-Name:").grid(row=0, column=0, sticky="w")
        self.category_name_var = tk.StringVar()
        tk.Entry(top_frame, textvariable=self.category_name_var, width=30).grid(row=0, column=1, sticky="w", padx=5)

        tk.Label(top_frame, text="Sheet-Name in Excel:").grid(row=1, column=0, sticky="w", pady=(5, 0))
        self.category_sheet_var = tk.StringVar()
        tk.Entry(top_frame, textvariable=self.category_sheet_var, width=30).grid(row=1, column=1, sticky="w", padx=5, pady=(5, 0))

        # Spaltenauswahl
        tk.Label(right_frame, text="Spalten auswählen:").pack(anchor="w", pady=(10, 0))

        columns_outer = tk.Frame(right_frame)
        columns_outer.pack(fill=tk.BOTH, expand=True)

        # Checkboxen für vordefinierte Spalten
        predefined_frame = tk.LabelFrame(columns_outer, text="Vordefinierte Spalten")
        predefined_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        # Canvas mit Scrollbar
        canvas = tk.Canvas(predefined_frame, width=200)
        scrollbar = ttk.Scrollbar(predefined_frame, orient="vertical", command=canvas.yview)
        self.predefined_checkboxes_frame = tk.Frame(canvas)

        self.predefined_checkboxes_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.predefined_checkboxes_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Checkbox-Variablen für vordefinierte Spalten
        self.column_checkbox_vars: Dict[str, tk.BooleanVar] = {}
        for col_name in PREDEFINED_COLUMNS.keys():
            var = tk.BooleanVar(value=False)
            self.column_checkbox_vars[col_name] = var
            cb = tk.Checkbutton(self.predefined_checkboxes_frame, text=col_name, variable=var)
            cb.pack(anchor="w")

        # Eigene Spalten
        custom_frame = tk.LabelFrame(columns_outer, text="Eigene Spalten (eine pro Zeile)")
        custom_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 5))

        self.custom_columns_text = tk.Text(custom_frame, width=20, height=15)
        self.custom_columns_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Spalten-Reihenfolge
        order_frame = tk.LabelFrame(columns_outer, text="Spalten-Reihenfolge")
        order_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))

        tk.Label(order_frame, text="(Auswahl → Button)", fg="gray", font=("TkDefaultFont", 8)).pack(anchor="w", padx=5, pady=(2, 0))
        
        list_frame = tk.Frame(order_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.column_order_listbox = tk.Listbox(list_frame, selectmode=tk.SINGLE, height=10)
        order_scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.column_order_listbox.yview)
        self.column_order_listbox.configure(yscrollcommand=order_scrollbar.set)
        
        self.column_order_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        order_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Buttons zum Verschieben
        order_btn_frame = tk.Frame(order_frame)
        order_btn_frame.pack(fill=tk.X, padx=5, pady=(0, 5))
        tk.Button(order_btn_frame, text="↑ Hoch", command=self.move_column_up, width=8).pack(side=tk.LEFT, padx=2)
        tk.Button(order_btn_frame, text="↓ Runter", command=self.move_column_down, width=8).pack(side=tk.LEFT, padx=2)
        tk.Button(order_btn_frame, text="🔄 Sync", command=self.sync_column_order, bg="#2196F3", fg="white", width=8).pack(side=tk.LEFT, padx=2)

        # Buttons
        btn_frame2 = tk.Frame(right_frame)
        btn_frame2.pack(fill=tk.X, pady=5)
        tk.Button(btn_frame2, text="Kategorie speichern", command=self.save_current_category).pack(side=tk.LEFT, padx=5)

    # ========== Global-Tab ==========
    def _build_global_tab(self, parent: tk.Frame) -> None:
        llm_frame = tk.LabelFrame(parent, text="LLM Einstellungen")
        llm_frame.pack(fill=tk.X, padx=5, pady=5)

        tk.Label(llm_frame, text="Provider:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.llm_provider_combo = ttk.Combobox(
            llm_frame,
            textvariable=self.llm_provider_var,
            state="readonly",
            width=25,
            values=["ollama", "google"],
        )
        self.llm_provider_combo.grid(row=0, column=1, sticky="ew", padx=5, pady=2)
        self.llm_provider_combo.bind("<<ComboboxSelected>>", self.on_llm_provider_change)

        tk.Label(llm_frame, text="Base URL:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        tk.Entry(llm_frame, textvariable=self.llm_base_url_var).grid(row=1, column=1, sticky="ew", padx=5, pady=2)

        tk.Label(llm_frame, text="Model:").grid(row=2, column=0, sticky="w", padx=5, pady=2)
        tk.Entry(llm_frame, textvariable=self.llm_model_var).grid(row=2, column=1, sticky="ew", padx=5, pady=2)

        tk.Label(llm_frame, text="API Key:").grid(row=3, column=0, sticky="w", padx=5, pady=2)
        tk.Entry(llm_frame, textvariable=self.llm_api_key_var, show="*").grid(
            row=3, column=1, sticky="ew", padx=5, pady=2
        )

        llm_frame.columnconfigure(1, weight=1)

        tk.Label(parent, text="Globaler Default-Prompt:").pack(anchor="w", pady=(10, 0), padx=5)
        self.default_prompt_text = tk.Text(parent, height=12)
        self.default_prompt_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        btn_frame = tk.Frame(parent)
        btn_frame.pack(fill=tk.X, pady=5)
        tk.Button(btn_frame, text="Konfiguration speichern", command=self.save_all).pack(side=tk.LEFT, padx=5)

    # ========== Vendor-Logik ==========
    def refresh_vendor_list(self) -> None:
        self.vendor_listbox.delete(0, tk.END)
        for vendor in self.cfg.get("vendors", []):
            self.vendor_listbox.insert(tk.END, vendor.get("name", "Unbenannt"))

    def refresh_vendor_category_choices(self) -> None:
        names = [cat.get("name", "") for cat in self.cfg.get("categories", [])]
        self.vendor_category_combo["values"] = names

    def current_vendor_index(self) -> int:
        return self._current_vendor_idx

    def on_vendor_select(self, event: Any) -> None:
        selection = self.vendor_listbox.curselection()
        if not selection:
            return
        idx = selection[0]
        self._current_vendor_idx = idx  # Index intern speichern

        try:
            vendor = self.cfg.get("vendors", [])[idx]
        except IndexError:
            return

        self.vendor_name_var.set(vendor.get("name", ""))
        self.vendor_category_var.set(vendor.get("category") or "")

        self.vendor_prompt_text.delete("1.0", tk.END)
        self.vendor_prompt_text.insert("1.0", vendor.get("prompt", "") or "")

        # Spalten-Prompts laden
        column_prompts = vendor.get("column_prompts", {}) or {}
        self._rebuild_column_prompts(column_prompts)

    def save_current_vendor(self) -> None:
        idx = self.current_vendor_index()
        if idx < 0:
            messagebox.showinfo("Info", "Bitte zuerst einen Vendor in der Liste auswählen oder hinzufügen.")
            return

        name = self.vendor_name_var.get().strip()
        prompt = self.vendor_prompt_text.get("1.0", tk.END).strip()
        category = self.vendor_category_var.get().strip() or None

        # Spalten-Prompts sammeln (kann jetzt Strings oder Objekte sein)
        column_prompts: Dict[str, Any] = {}
        for col_name, text_widget in self.column_prompt_entries.items():
            val = text_widget.get("1.0", tk.END).strip()
            if val:
                # Alte Struktur wird unterstützt (dict mit "prompt"), aber nur Prompt-Text speichern
                if isinstance(val, dict) and "prompt" in val:
                    column_prompts[col_name] = val["prompt"]
                else:
                    column_prompts[col_name] = val

        vendors = self.cfg.setdefault("vendors", [])
        if idx >= len(vendors):
            return
        vendors[idx]["name"] = name or "Unbenannt"
        vendors[idx]["prompt"] = prompt
        vendors[idx]["category"] = category
        vendors[idx]["column_prompts"] = column_prompts

        self.refresh_vendor_list()
        self.pipeline_refresh_vendor_choices()
        # Direkt in Datei speichern
        self._sync_llm_settings()
        self.cfg["default_prompt"] = self.default_prompt_text.get("1.0", tk.END).strip()
        save_config(self.cfg)
        messagebox.showinfo("Gespeichert", f"Vendor-Eintrag wurde aktualisiert und in {CONFIG_PATH} gespeichert.")

    def add_vendor(self) -> None:
        name = simpledialog.askstring("Neuer Vendor", "Name des Vendors:")
        if not name:
            return
        vendor = {
            "name": name.strip(),
            "patterns": [],
            "prompt": "",
            "category": None,
            "column_prompts": {},
        }
        self.cfg.setdefault("vendors", []).append(vendor)
        self.refresh_vendor_list()
        self.pipeline_refresh_vendor_choices()

        # Neuen Vendor automatisch auswählen
        new_idx = len(self.cfg.get("vendors", [])) - 1
        self._current_vendor_idx = new_idx  # Index intern setzen
        self.vendor_listbox.selection_clear(0, tk.END)
        self.vendor_listbox.selection_set(new_idx)
        self.vendor_listbox.see(new_idx)

        # Formular mit leeren/Default-Werten füllen
        self.vendor_name_var.set(name.strip())
        self.vendor_category_var.set("")
        self.vendor_prompt_text.delete("1.0", tk.END)
        self._rebuild_column_prompts()

    def delete_vendor(self) -> None:
        idx = self.current_vendor_index()
        if idx < 0:
            return
        if not messagebox.askyesno("Löschen", "Diesen Vendor wirklich löschen?"):
            return
        vendors = self.cfg.setdefault("vendors", [])
        if idx < len(vendors):
            vendors.pop(idx)
        self._current_vendor_idx = -1  # Auswahl zurücksetzen
        self.refresh_vendor_list()
        self.pipeline_refresh_vendor_choices()
        self.vendor_name_var.set("")
        self.vendor_category_var.set("")
        self.vendor_prompt_text.delete("1.0", tk.END)
        self._rebuild_column_prompts()

    # ========== Kategorien-Logik ==========
    def refresh_category_list(self) -> None:
        self.category_listbox.delete(0, tk.END)
        for cat in self.cfg.get("categories", []):
            self.category_listbox.insert(tk.END, cat.get("name", "Unbenannt"))

    def current_category_index(self) -> int:
        return self._current_category_idx

    def on_category_select(self, event: Any) -> None:
        selection = self.category_listbox.curselection()
        if not selection:
            return
        idx = selection[0]
        self._current_category_idx = idx  # Index intern speichern

        try:
            cat = self.cfg.get("categories", [])[idx]
        except IndexError:
            return

        self.category_name_var.set(cat.get("name", ""))
        self.category_sheet_var.set(cat.get("sheet_name", "") or "")

        cols: List[str] = cat.get("columns", []) or []

        # Checkboxen setzen
        for col_name, var in self.column_checkbox_vars.items():
            var.set(col_name in cols)

        # Custom columns (die nicht in PREDEFINED sind)
        custom_cols = [c for c in cols if c not in PREDEFINED_COLUMNS]
        self.custom_columns_text.delete("1.0", tk.END)
        self.custom_columns_text.insert("1.0", "\n".join(custom_cols))
        
        # Spalten-Reihenfolge Listbox aktualisieren
        self.column_order_listbox.delete(0, tk.END)
        for col in cols:
            self.column_order_listbox.insert(tk.END, col)

    def save_current_category(self) -> None:
        idx = self.current_category_index()
        if idx < 0:
            messagebox.showinfo("Info", "Bitte zuerst eine Kategorie auswählen oder hinzufügen.")
            return

        name = self.category_name_var.get().strip() or "Standard"
        sheet_name = self.category_sheet_var.get().strip() or name

        # Spalten aus der Order-Listbox holen (falls vorhanden)
        # Fallback: aus Checkboxen sammeln
        listbox_columns = list(self.column_order_listbox.get(0, tk.END))
        
        if listbox_columns:
            # Verwende die Reihenfolge aus der Listbox
            columns = listbox_columns
        else:
            # Fallback: Spalten aus Checkboxen sammeln
            columns: List[str] = []
            for col_name in PREDEFINED_COLUMNS.keys():
                if self.column_checkbox_vars[col_name].get():
                    columns.append(col_name)

            # Custom columns hinzufügen
            custom_raw = self.custom_columns_text.get("1.0", tk.END).strip()
            custom_cols = [c.strip() for c in custom_raw.splitlines() if c.strip()]
            for c in custom_cols:
                if c not in columns:
                    columns.append(c)

        cats = self.cfg.setdefault("categories", [])
        if idx >= len(cats):
            return
        cats[idx]["name"] = name
        cats[idx]["sheet_name"] = sheet_name
        cats[idx]["columns"] = columns

        self.refresh_category_list()
        self.refresh_vendor_category_choices()
        # Direkt in Datei speichern
        self._sync_llm_settings()
        self.cfg["default_prompt"] = self.default_prompt_text.get("1.0", tk.END).strip()
        save_config(self.cfg)
        messagebox.showinfo("Gespeichert", f"Kategorie wurde aktualisiert und in {CONFIG_PATH} gespeichert.")

    def add_category(self) -> None:
        name = simpledialog.askstring("Neue Kategorie", "Name der Kategorie:")
        if not name:
            return
        cat = {
            "name": name.strip(),
            "sheet_name": name.strip(),
            "columns": [],
        }
        self.cfg.setdefault("categories", []).append(cat)
        self.refresh_category_list()
        self.refresh_vendor_category_choices()

        # Neue Kategorie automatisch auswählen
        new_idx = len(self.cfg.get("categories", [])) - 1
        self._current_category_idx = new_idx  # Index intern setzen
        self.category_listbox.selection_clear(0, tk.END)
        self.category_listbox.selection_set(new_idx)
        self.category_listbox.see(new_idx)

        # Formular mit neuen Werten füllen
        self.category_name_var.set(name.strip())
        self.category_sheet_var.set(name.strip())
        for var in self.column_checkbox_vars.values():
            var.set(False)
        self.custom_columns_text.delete("1.0", tk.END)

    def delete_category(self) -> None:
        idx = self.current_category_index()
        if idx < 0:
            return
        if not messagebox.askyesno("Löschen", "Diese Kategorie wirklich löschen?"):
            return
        cats = self.cfg.setdefault("categories", [])
        if idx < len(cats):
            removed_name = cats[idx].get("name")
            cats.pop(idx)
            # Vendors, die diese Kategorie nutzen, auf None setzen
            for v in self.cfg.get("vendors", []):
                if v.get("category") == removed_name:
                    v["category"] = None
        self._current_category_idx = -1  # Auswahl zurücksetzen
        self.refresh_category_list()
        self.refresh_vendor_category_choices()
        self.category_name_var.set("")
        self.category_sheet_var.set("")
        for var in self.column_checkbox_vars.values():
            var.set(False)
        self.custom_columns_text.delete("1.0", tk.END)

    def import_excel_categories(self) -> None:
        """Lädt eine Excel-Datei und erstellt Kategorien aus den Sheets."""
        from tkinter import filedialog
        from openpyxl import load_workbook
        
        file_path = filedialog.askopenfilename(
            title="Excel-Tabelle auswählen",
            filetypes=[("Excel-Dateien", "*.xlsx *.xlsm"), ("Alle Dateien", "*.*")]
        )
        if not file_path:
            return
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            
            if not sheets:
                messagebox.showwarning("Warnung", "Die Excel-Datei enthält keine Sheets.")
                return
            
            # Dialog: Welche Sheets importieren?
            import_dialog = tk.Toplevel(self)
            import_dialog.title("Sheets importieren")
            import_dialog.geometry("500x400")
            import_dialog.transient(self)
            import_dialog.grab_set()
            
            tk.Label(import_dialog, text="Wähle die Sheets aus, die als Kategorien importiert werden sollen:", 
                     font=("TkDefaultFont", 10, "bold")).pack(padx=10, pady=10, anchor="w")
            
            # Scrollable frame für Checkboxen
            canvas_frame = tk.Frame(import_dialog)
            canvas_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            canvas = tk.Canvas(canvas_frame)
            scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
            checkboxes_frame = tk.Frame(canvas)
            
            checkboxes_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            canvas.create_window((0, 0), window=checkboxes_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            selected_sheets = {}
            for sheet_name in sheets:
                # Spalten aus diesem Sheet lesen
                ws = wb[sheet_name]
                columns = []
                if ws.max_row > 0:
                    for cell in ws[1]:
                        if cell.value:
                            col_name = str(cell.value).strip()
                            if col_name:
                                columns.append(col_name)
                
                var = tk.BooleanVar(value=True)
                frame = tk.Frame(checkboxes_frame)
                frame.pack(anchor="w", padx=5, pady=2, fill=tk.X)
                
                tk.Checkbutton(frame, text=f"{sheet_name}", variable=var, font=("TkDefaultFont", 9, "bold")).pack(side=tk.LEFT)
                tk.Label(frame, text=f"({len(columns)} Spalten)", fg="gray", font=("TkDefaultFont", 8)).pack(side=tk.LEFT, padx=5)
                
                selected_sheets[sheet_name] = (var, columns)
            
            def do_import():
                categories_to_add = []
                for sheet_name, (var, columns) in selected_sheets.items():
                    if var.get() and columns:
                        categories_to_add.append({
                            "name": sheet_name,
                            "sheet_name": sheet_name,
                            "columns": columns
                        })
                
                # Kategorien hinzufügen
                if categories_to_add:
                    existing = self.cfg.setdefault("categories", [])
                    existing_names = {c.get("name") for c in existing}
                    added_count = 0
                    
                    for new_cat in categories_to_add:
                        # Prüfen ob schon vorhanden
                        if new_cat["name"] not in existing_names:
                            existing.append(new_cat)
                            added_count += 1
                        else:
                            # Kategorie existiert schon - fragen ob überschreiben
                            if messagebox.askyesno("Kategorie existiert", 
                                                   f"Die Kategorie '{new_cat['name']}' existiert bereits.\n"
                                                   f"Überschreiben?"):
                                # Kategorie ersetzen
                                for i, cat in enumerate(existing):
                                    if cat.get("name") == new_cat["name"]:
                                        existing[i] = new_cat
                                        added_count += 1
                                        break
                    
                    if added_count > 0:
                        save_config(self.cfg)
                        self.refresh_category_list()
                        self.refresh_vendor_category_choices()
                        messagebox.showinfo("Erfolg", f"{added_count} Kategorie(n) importiert.")
                    else:
                        messagebox.showinfo("Info", "Keine neuen Kategorien importiert.")
                else:
                    messagebox.showwarning("Warnung", "Keine Sheets ausgewählt oder keine Spalten gefunden.")
                
                import_dialog.destroy()
            
            # Buttons
            btn_frame = tk.Frame(import_dialog)
            btn_frame.pack(fill=tk.X, padx=10, pady=10)
            tk.Button(btn_frame, text="Importieren", command=do_import, bg="#4CAF50", fg="white", width=15).pack(side=tk.LEFT, padx=5)
            tk.Button(btn_frame, text="Abbrechen", command=import_dialog.destroy, width=15).pack(side=tk.LEFT, padx=5)
            
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Laden der Excel-Datei:\n{str(e)}")

    def sync_column_order(self) -> None:
        """Synchronisiert die Spalten-Reihenfolge aus den Checkboxen."""
        # Sammle alle ausgewählten Spalten
        columns = []
        
        # Vordefinierte Spalten (in der Reihenfolge von PREDEFINED_COLUMNS)
        for col_name in PREDEFINED_COLUMNS.keys():
            if self.column_checkbox_vars[col_name].get():
                columns.append(col_name)
        
        # Custom Spalten
        custom_text = self.custom_columns_text.get("1.0", tk.END).strip()
        if custom_text:
            for line in custom_text.splitlines():
                col = line.strip()
                if col and col not in columns:
                    columns.append(col)
        
        # Listbox aktualisieren
        self.column_order_listbox.delete(0, tk.END)
        for col in columns:
            self.column_order_listbox.insert(tk.END, col)
    
    def move_column_up(self) -> None:
        """Bewegt die ausgewählte Spalte nach oben."""
        selection = self.column_order_listbox.curselection()
        if not selection or selection[0] == 0:
            return
        
        idx = selection[0]
        # Spalten aus Listbox holen
        columns = list(self.column_order_listbox.get(0, tk.END))
        
        # Tauschen
        columns[idx], columns[idx - 1] = columns[idx - 1], columns[idx]
        
        # Listbox aktualisieren
        self.column_order_listbox.delete(0, tk.END)
        for col in columns:
            self.column_order_listbox.insert(tk.END, col)
        
        # Auswahl beibehalten
        self.column_order_listbox.selection_set(idx - 1)
        self.column_order_listbox.see(idx - 1)
    
    def move_column_down(self) -> None:
        """Bewegt die ausgewählte Spalte nach unten."""
        selection = self.column_order_listbox.curselection()
        if not selection:
            return
        
        idx = selection[0]
        columns = list(self.column_order_listbox.get(0, tk.END))
        
        if idx >= len(columns) - 1:
            return
        
        # Tauschen
        columns[idx], columns[idx + 1] = columns[idx + 1], columns[idx]
        
        # Listbox aktualisieren
        self.column_order_listbox.delete(0, tk.END)
        for col in columns:
            self.column_order_listbox.insert(tk.END, col)
        
        # Auswahl beibehalten
        self.column_order_listbox.selection_set(idx + 1)
        self.column_order_listbox.see(idx + 1)

    # ========== Global ==========
    def load_default_prompt(self) -> None:
        self.default_prompt_text.delete("1.0", tk.END)
        self.default_prompt_text.insert("1.0", self.cfg.get("default_prompt", ""))

    def load_llm_settings(self) -> None:
        llm_cfg = self.cfg.get("llm", {})
        self.llm_provider_var.set(llm_cfg.get("provider", DEFAULT_LLM_CONFIG["provider"]))
        self.llm_base_url_var.set(llm_cfg.get("base_url", DEFAULT_LLM_CONFIG["base_url"]))
        self.llm_model_var.set(llm_cfg.get("model", DEFAULT_LLM_CONFIG["model"]))
        self.llm_api_key_var.set(llm_cfg.get("api_key", DEFAULT_LLM_CONFIG["api_key"]))

    def on_llm_provider_change(self, event: Any = None) -> None:
        provider = self.llm_provider_var.get()
        base_url = self.llm_base_url_var.get().strip()
        if provider == "google" and not base_url:
            self.llm_base_url_var.set("https://generativelanguage.googleapis.com/v1beta/openai")
        elif provider == "ollama" and not base_url:
            self.llm_base_url_var.set(DEFAULT_LLM_CONFIG["base_url"])

    def _sync_llm_settings(self) -> None:
        llm_cfg = self.cfg.setdefault("llm", {})
        llm_cfg["provider"] = self.llm_provider_var.get().strip() or DEFAULT_LLM_CONFIG["provider"]
        llm_cfg["base_url"] = self.llm_base_url_var.get().strip()
        llm_cfg["model"] = self.llm_model_var.get().strip()
        llm_cfg["api_key"] = self.llm_api_key_var.get().strip()
        if llm_cfg["provider"] == "google" and not llm_cfg["base_url"]:
            llm_cfg["base_url"] = "https://generativelanguage.googleapis.com/v1beta/openai"
        if llm_cfg["provider"] == "ollama" and not llm_cfg["base_url"]:
            llm_cfg["base_url"] = DEFAULT_LLM_CONFIG["base_url"]

    def save_all(self) -> None:
        self._sync_llm_settings()
        self.cfg["default_prompt"] = self.default_prompt_text.get("1.0", tk.END).strip()
        save_config(self.cfg)
        messagebox.showinfo("Gespeichert", f"Konfiguration in {CONFIG_PATH} gespeichert.")


if __name__ == "__main__":
    app = ConfigEditor()
    app.mainloop()
