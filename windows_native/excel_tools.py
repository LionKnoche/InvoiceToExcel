from typing import Callable, Optional

from openpyxl import load_workbook

try:
    from . import excel_store
except ImportError:
    import excel_store

LogFn = Callable[[str], None]


def lookup_excel_sheet(
    delivery_note: str,
    sheet_name: str,
    search_column: str,
    return_column: str,
    log_fn: Optional[LogFn] = None,
) -> str:
    """Lookup a value in an Excel sheet and return the matched column value."""
    try:
        excel_path = excel_store.EXCEL_PATH
        if not excel_path.exists():
            return f"Fehler: Excel-Datei nicht gefunden: {excel_path}"

        wb = load_workbook(excel_path, read_only=True)

        if sheet_name not in wb.sheetnames:
            return (
                f"Fehler: Sheet '{sheet_name}' nicht gefunden. "
                f"Verfuegbare Sheets: {', '.join(wb.sheetnames)}"
            )

        ws = wb[sheet_name]

        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append("")

        search_idx = None
        return_idx = None

        for idx, header in enumerate(headers, start=1):
            if header == search_column:
                search_idx = idx
            if header == return_column:
                return_idx = idx

        if search_idx is None:
            return (
                f"Fehler: Spalte '{search_column}' nicht gefunden. "
                f"Verfuegbare Spalten: {', '.join([h for h in headers if h])}"
            )

        if return_idx is None:
            return (
                f"Fehler: Spalte '{return_column}' nicht gefunden. "
                f"Verfuegbare Spalten: {', '.join([h for h in headers if h])}"
            )

        found_values = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            search_value = row[search_idx - 1].value
            if search_value and str(search_value).strip() == str(delivery_note).strip():
                return_value = row[return_idx - 1].value
                if return_value:
                    found_values.append(str(return_value))

        if found_values:
            return ", ".join(found_values)
        return ""
    except Exception as exc:
        if log_fn:
            log_fn(f"[Tool] Fehler bei Excel-Lookup: {exc}")
        return f"Fehler: {exc}"
