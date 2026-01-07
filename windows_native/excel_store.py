from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import date
from typing import List, Dict, Any, Optional

try:
    from . import config_store
except ImportError:
    import config_store

EXCEL_PATH = config_store.get_excel_path()
SHEET_NAME = "Rechnungen" 

# Standardspalten, falls kein konfiguriertes Schema verwendet wird.
COLUMNS = [ 
    "InvoiceId",
    "InvoiceNumber",
    "VendorName",
    "InvoiceDate",
    "DueDate",
    "GrossAmount",
    "TaxRate",
    "Currency",
    "Description",
    "FilePath",
]


def init_workbook(path: Path = EXCEL_PATH) -> None:
    """Erstellt die Excel-Datei mit Kopfzeile, falls sie noch nicht existiert (Standard-Schema)."""
    if path.exists(): # exists is a method of the Path class
        return # returns nothing if the file already exists; NONE was set as default return value in the function definition

    wb = Workbook() # Workbook is a class from the openpyxl library
    ws = wb.active # property that returns the active sheet
    ws.title = SHEET_NAME

    # Kopfzeile setzen (Standardschema)
    for idx, col_name in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=idx, value=col_name)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path) #after save the workbook is saved to the path


def _ensure_workbook(path: Path = EXCEL_PATH):
    """
    Hilfsfunktion: stellt sicher, dass Datei/SHEET existieren und lädt sie.
    Nutzt das Standardschema (SHEET_NAME/COLUMNS).
    """
    init_workbook(path)
    wb = load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        for idx, col_name in enumerate(COLUMNS, start=1):
            ws.cell(row=1, column=idx, value=col_name)
    ws = wb[SHEET_NAME]
    return wb, ws # returns a tuple of the workbook and the sheet


def ensure_workbook_for_schema(
    sheet_name: str,
    columns: List[str],
    path: Path = EXCEL_PATH,
):
    """
    Stellt sicher, dass die Arbeitsmappe und ein Sheet mit dem angegebenen
    Namen und Spaltenschema existiert.

    - Wenn die Datei noch nicht existiert, wird sie angelegt.
    - Wenn das Sheet neu ist, werden die Kopfzeilen aus `columns` geschrieben.
    """
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=idx, value=col_name)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return wb, ws

    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        for idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=idx, value=col_name)
    else:
        ws = wb[sheet_name]
    return wb, ws


def append_invoice( # this function is called by the mcp_excel_server.py file; the parameters are the values that are passed to the function
    invoice_number: str,
    vendor_name: str,
    invoice_date: date,
    due_date: Optional[date],
    gross_amount: float,
    tax_rate: float,
    currency: str,
    description: str,
    file_path: str,
    path: Path = EXCEL_PATH,
) -> int: 
    """
    Fügt eine neue Rechnung als Zeile hinzu.
    Gibt die neue InvoiceId zurück.
    """
    wb, ws = _ensure_workbook(path) # load the workbook and the sheet

    # Nächste Zeile = letzte belegte Zeile + 1
    next_row = ws.max_row + 1 # max row returns an integer; adds 1 to the last row

    # InvoiceId automatisch vergeben
    new_id = next_row - 1  # subtracts 1 from the last row to get the new id; it starts at 0, because the first row sets next_row to 1

    values = [
        new_id,
        invoice_number,
        vendor_name,
        invoice_date.isoformat(), #the isoformat method returns a string in the format YYYY-MM-DD
        due_date.isoformat() if due_date else None, # syntax is called a ternary operator; if due_date is not None, it returns the isoformat of the due date; otherwise it returns None
        float(gross_amount),
        float(tax_rate),
        currency,
        description,
        file_path,
    ]

    for col_index, value in enumerate(values, start=1): # enumerate is a function that is applied to an iterable
        ws.cell(row=next_row, column=col_index, value=value)

    wb.save(path)
    return new_id


def append_invoice_with_schema(
    data: Dict[str, Any],
    sheet_name: str,
    columns: List[str],
    path: Path = EXCEL_PATH,
) -> int:
    """
    Fügt eine Rechnung in ein bestimmtes Sheet mit einem konfigurierbaren
    Spaltenschema ein.

    Args:
        data: Dictionary mit allen Rechnungsdaten (beliebige Felder)
        sheet_name: Name des Excel-Sheets
        columns: Liste der Spaltennamen im Schema
        path: Pfad zur Excel-Datei

    Erwartet, dass `columns` die gewünschte Kopfzeilenreihenfolge enthält.
    Folgende logische Felder werden automatisch befüllt:
    - InvoiceId (autoincrement)
    - InvoiceNumber, VendorName, InvoiceDate, DueDate, GrossAmount, TaxRate, 
      Currency, Description, DeliveryNote, FilePath (aus data)
    """
    wb, ws = ensure_workbook_for_schema(sheet_name, columns, path=path)

    next_row = ws.max_row + 1
    new_id = next_row - 1  # Kopfzeile ist Zeile 1

    # Map von logischem Feldnamen zu Wert
    field_values: Dict[str, Any] = {
        "InvoiceId": new_id,
        "InvoiceNumber": data.get("InvoiceNumber", data.get("invoice_number", "")),
        "VendorName": data.get("VendorName", data.get("vendor_name", "")),
        "GrossAmount": float(data.get("GrossAmount", data.get("gross_amount", 0.0))),
        "TaxRate": float(data.get("TaxRate", data.get("tax_rate", 0.0))),
        "Currency": data.get("Currency", data.get("currency", "EUR")),
        "Description": data.get("Description", data.get("description", "")),
        "DeliveryNote": data.get("DeliveryNote", data.get("delivery_note", "")),
        "FilePath": data.get("FilePath", data.get("file_path", "")),
    }

    # Datumsfelder behandeln
    invoice_date = data.get("InvoiceDate", data.get("invoice_date"))
    if isinstance(invoice_date, date):
        field_values["InvoiceDate"] = invoice_date.isoformat()
    elif invoice_date:
        field_values["InvoiceDate"] = invoice_date
    else:
        field_values["InvoiceDate"] = None

    due_date = data.get("DueDate", data.get("due_date"))
    if isinstance(due_date, date):
        field_values["DueDate"] = due_date.isoformat()
    elif due_date:
        field_values["DueDate"] = due_date
    else:
        field_values["DueDate"] = None

    # Alle Spalten durchgehen und Werte schreiben
    for col_index, col_name in enumerate(columns, start=1):
        value = field_values.get(col_name)
        ws.cell(row=next_row, column=col_index, value=value)

    wb.save(path)
    return new_id


def read_invoices(path: Path = EXCEL_PATH) -> List[Dict[str, Any]]: 
    """Liest alle Rechnungen als Liste von Dictionaries."""
    wb, ws = _ensure_workbook(path)
    rows = list(ws.iter_rows(min_row=2, values_only=True)) 

    invoices: List[Dict[str, Any]] = []
    for row in rows:
        if all(value is None for value in row): # all is a function that is applied to an iterable; it returns True if all values are None; otherwise it returns False
            continue
        invoice = {col_name: row[idx] for idx, col_name in enumerate(COLUMNS)} # dict-comprehension
        invoices.append(invoice) # appends the invoice to the list;
        print(f"Processing row: {row}")  # curly braces mark where the variables are inserted

    return invoices


if __name__ == "__main__": # 
    # Einfacher Testlauf
    init_workbook()
    print("Alle Rechnungen:")
    for inv in read_invoices():
        print(inv) # prints the invoice as a dictionary; so inv is a dictionary with the column names as keys and the values as values

